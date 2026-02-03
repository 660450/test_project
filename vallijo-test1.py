from openpyxl import Workbook
from openpyxl.styles import PatternFill

def save_to_excel(data, filename="vallejo_model_color_latest.xlsx"):
    wb = Workbook(
    ws = wb.active
    ws.title = "Vallejo Model Color"

    # Headers
    ws.append(["Number", "Name", "HexColor", "ColorPreview"])

    for number, name, hex_color in data:
        # Voeg de rij toe
        ws.append([number, name, hex_color, ""])

        # Kleur de preview‑cel
        fill = PatternFill(start_color=hex_color.replace("#", ""),
                           end_color=hex_color.replace("#", ""),
                           fill_type="solid")

        preview_cell = ws[f"D{ws.max_row}"]
        preview_cell.fill = fill

    wb.save(filename)
    print(f"Excel opgeslagen als {filename}")
